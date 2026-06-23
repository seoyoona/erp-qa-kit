from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from erpqa.screen.audit import (
    _clean_key,
    _DYNAMIC_COL_RE,
    _resolve_binding_path,
    _screen_binding,
    run_screen_audit,
)
from erpqa.screen.extractors import (
    BackendModule,
    _is_shared_import,
    _role_from_proc,
    extract_frontend_feature,
    extract_spec_screen,
    resolve_backend_by_sps,
    resolve_backend_module,
    resolve_frontend_files,
)


class CleanKeyTests(unittest.TestCase):
    def test_alias_exact_then_case_insensitive(self):
        alias = {"iInQty": "in_qty"}
        alias_ci = {k.lower(): v for k, v in alias.items()}
        self.assertEqual(_clean_key("iInQty", alias, alias_ci), "in_qty")
        # spec wrote IYMF (uppercase) but backend alias is iYMF -> case-insensitive bridge
        self.assertEqual(_clean_key("IYMF", {}, {"iymf": "ym_from"}), "ym_from")

    def test_heuristic_strip_leading_i_any_case(self):
        self.assertEqual(_clean_key("iInYmd", {}, {}), "in_ymd")
        # IYM -> strip leading I -> "YM" -> "y_m"; matches FE "ym" via _fe_has flattening.
        self.assertEqual(_clean_key("IYM", {}, {}), "y_m")

    def test_dynamic_column_pattern(self):
        self.assertTrue(_DYNAMIC_COL_RE.match("spec13"))
        self.assertTrue(_DYNAMIC_COL_RE.match("day01"))
        self.assertFalse(_DYNAMIC_COL_RE.match("in_qty"))


class RoleTests(unittest.TestCase):
    def test_role_from_proc(self):
        self.assertEqual(_role_from_proc("MISPD.dbo.str_PDMaterialInput_IU"), "IU")
        self.assertEqual(_role_from_proc("str_PDMaterialInput_S"), "S")
        self.assertEqual(_role_from_proc("str_PDMaterialInput_D_20251217"), "D")

    def test_numbered_query_save_variants_map_to_base_role(self):
        # `_S2`/`_S4`/`_IU2` are numbered variants of the same role, not OTHER.
        self.assertEqual(_role_from_proc("str_PDPlanItemProduct_S2"), "S")
        self.assertEqual(_role_from_proc("str_PDPRGB01900_S4"), "S")
        self.assertEqual(_role_from_proc("str_Demo_IU2"), "IU")
        self.assertEqual(_role_from_proc("str_Demo_Report"), "OTHER")


class FrontendExtractTests(unittest.TestCase):
    def test_classifies_filter_and_column_schemas(self):
        with tempfile.TemporaryDirectory() as tmp:
            f = Path(tmp) / "model.ts"
            f.write_text(
                "export const zFooQueryParamsSchema = z.object({\n"
                "  plan_ym: z.string(),\n  item_no: z.string(),\n});\n"
                "export const zFooItemSchema = z.object({\n"
                "  in_qty: z.number(),\n  lot_no: z.string(),\n});\n",
                encoding="utf-8",
            )
            fe = extract_frontend_feature([f])
            self.assertIn("plan_ym", fe.filter_fields)
            self.assertIn("item_no", fe.filter_fields)
            self.assertIn("in_qty", fe.column_fields)
            self.assertNotIn("in_qty", fe.filter_fields)


class SpecExtractTests(unittest.TestCase):
    def test_extracts_sp_params_and_roles(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "PDT-OSC-001M_demo.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "화면 ID"
            ws["B1"] = "PDT-OSC-001M"
            ws["A3"] = "exec MISPD.dbo.str_Demo_IU @iInYmd='20250101',@iInQty=10,@iLotNo='L1'"
            ws["A4"] = "exec MISPD.dbo.str_Demo_S @iPlanYm='202501'"
            wb.save(path)
            spec = extract_spec_screen(path, "PDT-OSC-001M")
            self.assertEqual(spec.params_for_role("IU"), ["iInYmd", "iInQty", "iLotNo"])
            self.assertEqual(spec.params_for_role("S"), ["iPlanYm"])


def _mk_backend_dir(root, name, body="x = 1\n"):
    d = root / name
    d.mkdir(parents=True)
    (d / "service.py").write_text(body, encoding="utf-8")
    return d


class ResolveBackendModuleTests(unittest.TestCase):
    def test_exact_normalized_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            _mk_backend_dir(root, "pdt_osc_001m")
            self.assertEqual(resolve_backend_module(root, "PDT-OSC-001M").name, "pdt_osc_001m")

    def test_transposition_binds_prg_to_pgr(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            _mk_backend_dir(root, "pdt_pgr_003m")
            self.assertEqual(resolve_backend_module(root, "PDT_PRG_003M").name, "pdt_pgr_003m")

    def test_digits_do_not_collide_003_vs_004(self):
        # The over-binding bug: 003 must not resolve to a 004-only module.
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            _mk_backend_dir(root, "pdt_pgr_004m")
            self.assertIsNone(resolve_backend_module(root, "PDT_PRG_003M"))

    def test_digit_anagram_does_not_collide_010_vs_001(self):
        # Digits keep their sequence: 010 must not transposition-match a 001 module
        # (sorted-multiset would have wrongly treated {0,1,0} == {0,0,1}).
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            _mk_backend_dir(root, "pdt_pgr_001m")
            self.assertIsNone(resolve_backend_module(root, "PDT_PRG_010M"))
            # but the real 010 module still binds via transposition
            _mk_backend_dir(root, "pdt_pgr_010m")
            self.assertEqual(resolve_backend_module(root, "PDT_PRG_010M").name, "pdt_pgr_010m")


class ResolveBackendBySpsTests(unittest.TestCase):
    def test_unique_sp_outweighs_shared_generic_sp(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            # Both modules contain the generic shared SP; only one has the unique SP.
            _mk_backend_dir(root, "mod_a", "q = 'str_Common_S'\n")
            _mk_backend_dir(root, "mod_b", "q = 'str_Common_S'\nu = 'str_Unique003_S'\n")
            picked = resolve_backend_by_sps(root, ["str_Common_S", "str_Unique003_S"])
            self.assertEqual(picked.name, "mod_b")


class ScreenBindingTests(unittest.TestCase):
    def test_screen_binding_matches_on_normalized_id(self):
        manifest = {"screen_bindings": {"PDT_PRG_003M": {"backend": "b", "frontend": "f"}}}
        self.assertEqual(_screen_binding(manifest, "PDT-PRG-003M"), {"backend": "b", "frontend": "f"})
        self.assertEqual(_screen_binding(manifest, "PDT_XXX_999M"), {})

    def test_resolve_binding_path_prefers_root_then_project(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "backend"
            (root / "pdt_pgr_003m").mkdir(parents=True)
            got = _resolve_binding_path("pdt_pgr_003m", root, Path(tmp))
            self.assertEqual(got, (root / "pdt_pgr_003m").resolve())
            self.assertIsNone(_resolve_binding_path("nope", root, Path(tmp)))
            self.assertIsNone(_resolve_binding_path(None, root))


class CopyStepRuleTests(unittest.TestCase):
    def test_is_shared_import(self):
        self.assertTrue(_is_shared_import("@/models/raw-material"))
        self.assertTrue(_is_shared_import("@/adapters/api/client"))
        self.assertTrue(_is_shared_import("../../models/foo"))
        self.assertFalse(_is_shared_import("./localThing"))
        self.assertFalse(_is_shared_import("react"))

    def _backend_with_sp(self, tmp, sp="str_Demo"):
        d = Path(tmp) / "be"
        d.mkdir()
        (d / "svc.py").write_text(f"q = '{sp}_S'\n", encoding="utf-8")
        return BackendModule(module_dir=str(d))

    def test_missing_shared_model_is_detected(self):
        with tempfile.TemporaryDirectory() as tmp:
            fe = Path(tmp) / "fe"
            (fe / "feature").mkdir(parents=True)
            # seed references the SP base, and imports a shared model not in the slice
            (fe / "feature" / "Controller.tsx").write_text(
                "import { z } from 'zod';\n"
                "import { schema } from '@/models/raw-material';\n"
                "const q = 'str_Demo_S';\n", encoding="utf-8")
            backend = self._backend_with_sp(tmp)
            files, anchor, missing = resolve_frontend_files(fe, backend, "PDT-X-001M")
            self.assertIn("@/models/raw-material", missing)

    def test_present_shared_model_is_not_flagged(self):
        with tempfile.TemporaryDirectory() as tmp:
            fe = Path(tmp) / "fe"
            (fe / "feature").mkdir(parents=True)
            (fe / "models").mkdir(parents=True)
            (fe / "models" / "raw-material.ts").write_text(
                "export const schema = z.object({ a: z.string() });\n", encoding="utf-8")
            (fe / "feature" / "Controller.tsx").write_text(
                "import { schema } from '@/models/raw-material';\n"
                "const q = 'str_Demo_S';\n", encoding="utf-8")
            backend = self._backend_with_sp(tmp)
            files, anchor, missing = resolve_frontend_files(fe, backend, "PDT-X-001M")
            self.assertEqual(missing, set())


class _Fixture:
    """Build a minimal project where one screen-id does NOT map to any module dir,
    forcing the sp-fallback path (the unreliable resolution)."""

    def __init__(self, tmp, screen_id="PDT-PRG-003M", bind=False, name_suffix="progress"):
        from openpyxl import Workbook
        import yaml

        self.project = Path(tmp)
        mod = "PDT"
        qa = self.project / "qa-context" / "modules" / mod
        qa.mkdir(parents=True)
        backend_root = self.project / "extracted" / mod / "backend"
        # module dir name shares no id-code with the screen -> id-code resolve fails.
        bdir = backend_root / "legacy_progress_module"
        bdir.mkdir(parents=True)
        (bdir / "svc.py").write_text("q = 'str_PgrCommon_S'\n", encoding="utf-8")
        fe_root = self.project / "extracted" / mod / "frontend"
        fe_root.mkdir(parents=True)
        (fe_root / "model.ts").write_text(
            "export const zItemSchema = z.object({ plan_ym: z.string() });\n", encoding="utf-8"
        )
        spec_root = self.project / "extracted" / mod / "spec"
        spec_root.mkdir(parents=True)
        wb = Workbook()
        ws = wb.active
        ws["A1"] = screen_id
        ws["A2"] = "exec MISPD.dbo.str_PgrCommon_S @iPlanYm='202501'"
        wb.save(spec_root / f"{screen_id}_{name_suffix}.xlsx")
        manifest = {"module": mod, "source_roots": {
            "spec": [f"extracted/{mod}/spec"], "frontend": [f"extracted/{mod}/frontend"],
            "backend": [f"extracted/{mod}/backend"]}}
        if bind:
            manifest["screen_bindings"] = {screen_id: {
                "backend": f"extracted/{mod}/backend/legacy_progress_module",
                "frontend": f"extracted/{mod}/frontend"}}
        (qa / "module_manifest.yaml").write_text(yaml.safe_dump(manifest), encoding="utf-8")
        self.module = mod
        self.screen_id = screen_id


class ResolutionHonestyTests(unittest.TestCase):
    def test_sp_fallback_is_flagged_low_confidence(self):
        with tempfile.TemporaryDirectory() as tmp:
            fx = _Fixture(tmp, bind=False)
            r = run_screen_audit(fx.project, fx.module, fx.screen_id)
            self.assertEqual(r["resolution"]["method"], "sp-fallback")
            self.assertTrue(r["resolution"]["uncertain"])
            types = [f["mismatch_type"] for f in r["findings"]]
            self.assertIn("ScreenResolutionLowConfidence", types)
            self.assertTrue(all(f.get("confidence") == "low" for f in r["findings"]))

    def test_explicit_binding_is_deterministic_and_trusted(self):
        with tempfile.TemporaryDirectory() as tmp:
            fx = _Fixture(tmp, bind=True)
            r = run_screen_audit(fx.project, fx.module, fx.screen_id)
            self.assertEqual(r["resolution"]["method"], "explicit")
            self.assertFalse(r["resolution"]["uncertain"])
            types = [f["mismatch_type"] for f in r["findings"]]
            self.assertNotIn("ScreenResolutionLowConfidence", types)


class SaveSpecCoverageTests(unittest.TestCase):
    def test_save_screen_without_save_sp_is_flagged(self):
        # Spec name says 등록 but carries only an S proc -> save dimension can't be
        # evaluated; must surface SaveSpecExampleMissing, not silent CLEAN.
        with tempfile.TemporaryDirectory() as tmp:
            fx = _Fixture(tmp, bind=True, name_suffix="생산계획등록")
            r = run_screen_audit(fx.project, fx.module, fx.screen_id)
            types = [f["mismatch_type"] for f in r["findings"]]
            self.assertIn("SaveSpecExampleMissing", types)

    def test_query_only_screen_is_not_flagged(self):
        with tempfile.TemporaryDirectory() as tmp:
            fx = _Fixture(tmp, bind=True, name_suffix="진행현황")
            r = run_screen_audit(fx.project, fx.module, fx.screen_id)
            types = [f["mismatch_type"] for f in r["findings"]]
            self.assertNotIn("SaveSpecExampleMissing", types)


class DiscoverScreenIdsTests(unittest.TestCase):
    def test_discovers_id_coded_specs_and_skips_others(self):
        from erpqa.screen.extractors import discover_screen_ids
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for n in ("PDT-OSC-001M_원부자재입고등록.xlsx", "PDT_PRG_003M_반제품진행.xlsx",
                      "PDT_MG_001M_생산계획현황.xlsx", "PDT_기준정보_공정등록.xlsx",
                      "PDT_생산데이터_금형관리.xlsx", "PDT_PRG_003M_반제품_미사용.xlsx"):
                (root / n).write_bytes(b"")
            ids = discover_screen_ids(root)
            # underscore/dash both normalize to dash form; dup PRG-003M collapses
            self.assertEqual(ids, ["PDT-MG-001M", "PDT-OSC-001M", "PDT-PRG-003M"])


class VisionLabelTests(unittest.TestCase):
    def test_is_label_keeps_korean_captions_drops_data(self):
        from erpqa.screen.vision import is_label
        for good in ("입고년월", "품목코드", "거래처", "공급가액"):
            self.assertTrue(is_label(good), good)
        for bad in (
            "A01-013",            # code
            "2025년 3월",          # has digits
            "(주)다인에스티에스",   # company marker
            "다인입고 (컬링)",      # parenthesised data
            "5",                  # number
            "OK",                 # no hangul
            "가",                 # too short / single hangul
        ):
            self.assertFalse(is_label(bad), bad)

    def test_extract_spec_labels_uses_injected_ocr_and_filters(self):
        # Build a tiny xlsx with an embedded PNG; inject a fake OCR so the test does
        # not depend on the platform Vision engine.
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from erpqa.screen import vision

        with tempfile.TemporaryDirectory() as tmp:
            png = Path(tmp) / "shot.png"
            # 1x1 PNG
            png.write_bytes(bytes.fromhex(
                "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
                "890000000d4944415478da6360000002000154a24f3f0000000049454e44ae426082"
            ))
            wb = Workbook(); ws = wb.active
            ws.add_image(XLImage(str(png)), "A1")
            xlsx = Path(tmp) / "PDT-X-001M.xlsx"; wb.save(xlsx)

            fake_ocr = lambda p: ["입고년월", "품목코드", "A01-013", "(주)데모", "5"]
            labels, meta = vision.extract_spec_labels(xlsx, Path(tmp) / "imgs", ocr=fake_ocr)
            self.assertEqual(labels, ["입고년월", "품목코드"])
            self.assertEqual(meta["ocr_engine"], "macos-vision")
            # openpyxl renames embedded media on save; just assert a PNG was read.
            self.assertTrue(meta["images_read"])
            self.assertTrue(meta["images_read"][0].endswith(".png"))


if __name__ == "__main__":
    unittest.main()
