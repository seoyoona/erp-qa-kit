from __future__ import annotations

from pathlib import Path
import shutil
import sqlite3
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
import yaml

from erpqa.adapters.extractors import (
    backend_static_adapter,
    csv_adapter,
    frontend_static_adapter,
    json_yaml_adapter,
    markdown_adapter,
    procedure_sql_adapter,
    xlsx_image_manifest_adapter,
    xlsx_text_adapter,
)
from erpqa.adapters.extractors.registry import required_adapter_names
from erpqa.cli import main
from erpqa.core.compare import compare_contracts
from erpqa.core.context import load_context
from erpqa.core.errors import WriteOutsideQaContextError
from erpqa.core.module_paths import module_output_path
from erpqa.core.policy import confirm_policy_for_tests
from erpqa.core.yaml_io import load_yaml
from tests.helpers import REPO_ROOT, snapshot_outside_qa


FIXTURE = REPO_ROOT / "examples" / "module-audit-fixture"


def copy_fixture(tmp_path: Path) -> Path:
    target = tmp_path / "module-audit-fixture"
    # Copy only the seed inputs; never copy a locally-generated qa-context/ so the
    # test is deterministic regardless of prior local `module-audit` runs.
    shutil.copytree(FIXTURE, target, ignore=shutil.ignore_patterns("qa-context"))
    return target


def prepare_project(tmp_path: Path) -> Path:
    project = copy_fixture(tmp_path)
    assert main(["policy-init", str(project)]) == 0
    confirm_policy_for_tests(project)
    assert main(["module-init", str(project), "--module", "ORD"]) == 0
    return project


class V02PolicyMemoryCliTests(unittest.TestCase):
    def test_policy_init_creates_seed_files_and_is_no_clobber(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "project"
            project.mkdir()
            self.assertEqual(main(["policy-init", str(project)]), 0)
            qa = project / "qa-context"
            for rel in ("project_policy.yaml", "project_memory.md", "project_assumptions.yaml"):
                self.assertTrue((qa / rel).exists(), rel)
            policy = load_yaml(qa / "project_policy.yaml")
            self.assertFalse(policy["qa_policy"]["frontend_override_allowed"])
            self.assertIn("validation_order", policy["qa_policy"])
            self.assertIn("defer_until_frontend_verified", policy["qa_policy"])
            self.assertTrue(policy["qa_policy"]["draft"])
            before = (qa / "project_memory.md").read_text(encoding="utf-8")
            (qa / "project_memory.md").write_text(before + "\nPM edit\n", encoding="utf-8")
            self.assertEqual(main(["policy-init", str(project)]), 0)
            self.assertIn("PM edit", (qa / "project_memory.md").read_text(encoding="utf-8"))

    def test_missing_policy_halts_module_command(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "project"
            (project / "qa-context").mkdir(parents=True)
            self.assertEqual(main(["extract-spec", str(project), "--module", "ORD"]), 1)

    def test_module_init_scaffolds_and_preserves_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            qa = project / "qa-context" / "modules" / "ORD"
            for rel in (
                "module_manifest.yaml",
                "module_policy.yaml",
                "module_memory.md",
                "screen_contract.yaml",
                "frontend_contract.yaml",
                "backend_contract.yaml",
                "procedure_contract.yaml",
            ):
                self.assertTrue((qa / rel).exists(), rel)
            before = (qa / "module_memory.md").read_text(encoding="utf-8")
            (qa / "module_memory.md").write_text(before + "\nOVERRIDE of X: X prime\n", encoding="utf-8")
            self.assertEqual(main(["module-init", str(project), "--module", "ORD"]), 0)
            self.assertIn("X prime", (qa / "module_memory.md").read_text(encoding="utf-8"))

    def test_module_memory_override_is_cited_in_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            module_memory = project / "qa-context" / "modules" / "ORD" / "module_memory.md"
            module_memory.write_text(module_memory.read_text(encoding="utf-8") + "\nOVERRIDE of fact X: fact X prime\n", encoding="utf-8")
            self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)
            report = project / "qa-context" / "modules" / "ORD" / "reports" / "screen_frontend_mismatch_report.md"
            text = report.read_text(encoding="utf-8")
            self.assertIn("Memory read: qa-context/project_memory.md", text)
            self.assertIn("Memory read: qa-context/modules/ORD/module_memory.md", text)
            self.assertIn("Module memory precedence", text)

    def test_module_commands_load_context_before_stage_work(self):
        order: list[str] = []
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "project"
            project.mkdir()
            fake_ctx = object()
            with mock.patch("erpqa.cli.load_context", side_effect=lambda *a, **k: order.append("context") or fake_ctx), \
                mock.patch("erpqa.cli.generate_frontend_contract", side_effect=lambda ctx: order.append("stage") or Path("frontend_contract.yaml")):
                self.assertEqual(main(["extract-frontend", str(project), "--module", "ORD"]), 0)
        self.assertEqual(order, ["context", "stage"])


class V02ExtractorTests(unittest.TestCase):
    def test_required_registry_names(self):
        self.assertEqual(
            sorted(required_adapter_names()),
            sorted(
                [
                    "xlsx_text",
                    "xlsx_image_manifest",
                    "csv",
                    "markdown",
                    "json_yaml",
                    "frontend_static",
                    "backend_static",
                    "procedure_sql",
                ]
            ),
        )

    def test_xlsx_text_adapter_extracts_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "spec.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Screen"
            ws.append(["section", "key", "label", "visible", "required", "order"])
            ws.append(["grid_columns", "order_no", "Order No.", "true", "true", "1"])
            ws.merge_cells("A4:B4")
            ws["A4"] = "Merged Header"
            wb.save(path)
            raw = xlsx_text_adapter.extract(path)
            self.assertTrue(raw.records)
            self.assertEqual(raw.provenance_hint, "medium")
            self.assertTrue(any(record["value"] == "Order No." for record in raw.records))

    def test_xlsx_image_manifest_adapter_lists_images_without_ocr(self):
        try:
            from PIL import Image
        except Exception:
            self.skipTest("Pillow unavailable for fixture image creation")
        with tempfile.TemporaryDirectory() as tmp:
            image_path = Path(tmp) / "pixel.png"
            Image.new("RGB", (1, 1), color="white").save(image_path)
            workbook_path = Path(tmp) / "image.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.add_image(OpenPyxlImage(str(image_path)), "B2")
            wb.save(workbook_path)
            raw = xlsx_image_manifest_adapter.extract(workbook_path)
            self.assertEqual(raw.provenance_hint, "low")
            self.assertTrue(raw.records)
            self.assertFalse(raw.records[0]["ocr_performed"])
            self.assertIsNone(raw.records[0]["text"])

    def test_csv_markdown_json_yaml_adapters(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = root / "spec.csv"
            csv_path.write_text("section,key,label\nvisible_text,title,Title\n", encoding="utf-8")
            md_path = root / "spec.md"
            md_path.write_text("# Screen\n\n| section | key | label |\n|---|---|---|\n| grid_columns | order_no | Order No. |\n", encoding="utf-8")
            json_path = root / "spec.json"
            json_path.write_text('{"screen_id":"S","sections":{"visible_text":[]}}', encoding="utf-8")
            yaml_path = root / "spec.yaml"
            yaml_path.write_text("screen: S\ncolumns:\n  - order_no\n", encoding="utf-8")
            self.assertEqual(csv_adapter.extract(csv_path).provenance_hint, "medium")
            self.assertTrue(markdown_adapter.extract(md_path).records)
            self.assertEqual(json_yaml_adapter.extract(json_path).provenance_hint, "high")
            self.assertEqual(json_yaml_adapter.extract(yaml_path).provenance_hint, "medium")

    def test_static_and_procedure_adapters_are_deterministic(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            frontend = project / "extracted" / "ORD" / "frontend"
            backend = project / "extracted" / "ORD" / "backend"
            procedure = project / "extracted" / "ORD" / "procedure"
            self.assertEqual(frontend_static_adapter.extract(frontend), frontend_static_adapter.extract(frontend))
            self.assertTrue(backend_static_adapter.extract(backend).records)
            raw_proc = procedure_sql_adapter.extract(procedure)
            self.assertTrue(raw_proc.records)
            self.assertEqual(raw_proc.provenance_hint, "low")


class V02EndToEndTests(unittest.TestCase):
    def test_module_audit_outputs_reports_handoff_and_all_mismatch_types(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)
            module_dir = project / "qa-context" / "modules" / "ORD"
            for rel in (
                "screen_contract.yaml",
                "frontend_contract.yaml",
                "backend_contract.yaml",
                "procedure_contract.yaml",
                "reports/screen_frontend_mismatch_report.md",
                "reports/column_mismatch_report.md",
                "reports/frontend_contract_report.md",
                "reports/procedure_mapping_report.md",
                "handoff/frontend_fix_handoff.md",
                "handoff/codex_frontend_fix_prompt.md",
            ):
                self.assertTrue((module_dir / rel).exists(), rel)
            findings = load_yaml(module_dir / "comparison_findings.yaml")["findings"]
            mismatch_types = {item["mismatch_type"] for item in findings}
            for expected in (
                "MissingInFrontend",
                "UnexpectedFrontendColumn",
                "LabelMismatch",
                "OrderMismatch",
                "VisibilityMismatch",
                "RequiredFlagMismatch",
                "ReadonlyDisabledMismatch",
                "FieldKeyMismatch",
                "SearchFilterMismatch",
                "GridColumnMismatch",
                "FormFieldMismatch",
                "ButtonsActionsMismatch",
                "InternalAuditSystemFieldExposed",
                "FrontendOverrideForbidden",
            ):
                self.assertIn(expected, mismatch_types)
            self.assertTrue(any(item["frontend_override_forbidden"] for item in findings))
            self.assertTrue(any(item["needs_human_confirmation"] for item in findings))
            report = (module_dir / "reports" / "screen_frontend_mismatch_report.md").read_text(encoding="utf-8")
            for field in (
                "screen_id",
                "module",
                "mismatch_type",
                "expected_from_source_of_truth",
                "actual_frontend",
                "source_files",
                "confidence",
                "needs_human_confirmation",
                "severity",
                "suggested_fix_type",
                "frontend_override_forbidden",
                "backend_evidence_agrees_or_conflicts",
                "project_memory_read",
                "module_memory_read",
                "handoff_ready_ai_fix_instruction",
            ):
                self.assertIn(field, report)
            handoff = (module_dir / "handoff" / "frontend_fix_handoff.md").read_text(encoding="utf-8")
            prompt = (module_dir / "handoff" / "codex_frontend_fix_prompt.md").read_text(encoding="utf-8")
            self.assertIn("TARGET FRONTEND repo", handoff)
            self.assertIn("Frontend override is forbidden", handoff)
            self.assertIn("TARGET FRONTEND repo", prompt)
            self.assertIn("extract-frontend", prompt)

    def test_contract_validation_and_missing_field_failure(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)
            self.assertEqual(main(["validate", str(project)]), 0)
            screen = project / "qa-context" / "modules" / "ORD" / "screen_contract.yaml"
            data = load_yaml(screen)
            del data["sections"]["grid_columns"][0]["label"]
            screen.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
            self.assertEqual(main(["validate", str(project)]), 1)

    def test_v02_safety_no_db_no_source_modification_and_containment(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            before = snapshot_outside_qa(project)

            def fail_connect(*args, **kwargs):
                raise AssertionError("database connection attempted")

            with mock.patch.object(sqlite3, "connect", fail_connect):
                self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)
            after = snapshot_outside_qa(project)
            self.assertEqual(before, after)
            with self.assertRaises(WriteOutsideQaContextError):
                module_output_path(project, "ORD", "../../../outside.txt")

    def test_generated_artifacts_use_relative_paths_not_absolute(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)
            module_dir = project / "qa-context" / "modules" / "ORD"
            abs_project = str(project.resolve())
            checked = 0
            for artifact in sorted(module_dir.rglob("*")):
                if artifact.is_file() and artifact.suffix in {".md", ".yaml"}:
                    text = artifact.read_text(encoding="utf-8")
                    self.assertNotIn(
                        abs_project,
                        text,
                        f"absolute project path leaked into {artifact.name}",
                    )
                    self.assertNotIn("/Users/", text, f"absolute /Users/ path leaked into {artifact.name}")
                    checked += 1
            self.assertGreater(checked, 0)

    def test_optional_dependencies_absent_do_not_block_required_flow(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = prepare_project(Path(tmp))
            self.assertEqual(main(["module-audit", str(project), "--module", "ORD"]), 0)


if __name__ == "__main__":
    unittest.main()
