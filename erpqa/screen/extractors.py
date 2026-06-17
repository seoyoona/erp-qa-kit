from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path


# --------------------------------------------------------------------------- #
# Data containers
# --------------------------------------------------------------------------- #
@dataclass
class SpecScreen:
    screen_id: str
    source_file: str
    sps: list[dict] = field(default_factory=list)  # {proc, role, params:[...]}

    def params_for_role(self, role: str) -> list[str]:
        out: list[str] = []
        for sp in self.sps:
            if sp["role"] == role:
                for p in sp["params"]:
                    if p not in out:
                        out.append(p)
        return out


@dataclass
class BackendModule:
    module_dir: str
    alias_to_clean: dict[str, str] = field(default_factory=dict)   # iInQty -> in_qty / INQTY -> in_qty
    clean_keys: set[str] = field(default_factory=set)
    sp_param_to_clean: dict[str, str] = field(default_factory=dict)  # iInYmd -> in_ymd (from service dicts)
    sp_param_hardcoded: dict[str, str] = field(default_factory=dict)  # iItemNo -> "" (hardcoded in service)
    implemented_roles: set[str] = field(default_factory=set)       # {S, IU, D}
    param_keys: set[str] = field(default_factory=set)              # from serialization_alias (proc inputs)
    result_columns: set[str] = field(default_factory=set)          # from validation_alias (S result rows)


@dataclass
class FrontendFeature:
    feature_dir: str
    files: list[str] = field(default_factory=list)
    zod_fields: set[str] = field(default_factory=set)
    filter_fields: set[str] = field(default_factory=set)           # from *QueryParams*/*Filter* schemas
    column_fields: set[str] = field(default_factory=set)           # from *Item*/*Row*/*Detail* schemas + grid defs
    mutations: set[str] = field(default_factory=set)               # {create, update, delete}


# --------------------------------------------------------------------------- #
# Spec (xlsx 화면IO) extractor
# --------------------------------------------------------------------------- #
_EXEC_RE = re.compile(r"exec\s+([\w.]+)\s*(.*)", re.IGNORECASE)
_PARAM_RE = re.compile(r"@(\w+)\s*=")
_SCREEN_ID_RE = re.compile(r"\bPDT[-_][A-Z0-9]+[-_]?\d+M\b", re.IGNORECASE)


_ROLE_SUFFIX_RE = re.compile(r"^(IU|S|I|U|D)\d*$")  # S, S2, S4, IU, IU2 ...


def _role_from_proc(proc: str) -> str:
    base = proc.rsplit(".", 1)[-1]
    base = re.sub(r"_\d{6,}$", "", base)  # strip dated suffix e.g. _20251217
    suffix = base.rsplit("_", 1)[-1].upper()
    # Numbered query/save variants (`_S2`, `_S4`, `_IU2`) carry the same role as
    # their base — only the trailing digits differ — so map them back, otherwise
    # their params (filters / save fields) are silently lost as OTHER.
    m = _ROLE_SUFFIX_RE.match(suffix)
    return m.group(1) if m else "OTHER"


def extract_spec_screen(xlsx_path: Path, screen_id: str | None = None) -> SpecScreen:
    from openpyxl import load_workbook  # required dep

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    texts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip():
                    texts.append(str(cell).strip())
    wb.close()

    sps: list[dict] = []
    found_id = screen_id
    for text in texts:
        if found_id is None:
            m = _SCREEN_ID_RE.search(text)
            if m:
                found_id = m.group(0).upper()
        for line in text.splitlines():
            em = _EXEC_RE.search(line)
            if em:
                proc, args = em.group(1), em.group(2)
                params = _PARAM_RE.findall(args)
                sps.append({"proc": proc, "role": _role_from_proc(proc), "params": params})

    return SpecScreen(
        screen_id=found_id or (screen_id or xlsx_path.stem),
        source_file=xlsx_path.name,
        sps=sps,
    )


# --------------------------------------------------------------------------- #
# Backend (python module) extractor
# --------------------------------------------------------------------------- #
_SER_ALIAS_RE = re.compile(r"(\w+)\s*:[^=\n]*=\s*Field\([^)]*serialization_alias\s*=\s*[\"'](\w+)[\"']")
_VAL_ALIAS_RE = re.compile(r"(\w+)\s*:[^=\n]*=\s*Field\([^)]*validation_alias\s*=\s*[\"'](\w+)[\"']")
_DICT_MAP_RE = re.compile(r"[\"'](i[A-Za-z]\w*)[\"']\s*:\s*payload\.(\w+)")
_DICT_HARD_RE = re.compile(r"[\"'](i[A-Za-z]\w*)[\"']\s*:\s*[\"']([^\"']*)[\"']")
_SP_TOKEN_RE = re.compile(r"str_\w+?_(IU|S|D)\b")


def extract_backend_module(module_dir: Path) -> BackendModule:
    bm = BackendModule(module_dir=str(module_dir))
    for py in sorted(module_dir.rglob("*.py")):
        src = py.read_text(encoding="utf-8", errors="replace")
        for clean, alias in _SER_ALIAS_RE.findall(src):
            bm.alias_to_clean[alias] = clean
            bm.clean_keys.add(clean)
            bm.param_keys.add(clean)
        for clean, alias in _VAL_ALIAS_RE.findall(src):
            bm.alias_to_clean[alias] = clean
            bm.clean_keys.add(clean)
            bm.result_columns.add(clean)
        for alias, clean in _DICT_MAP_RE.findall(src):
            bm.sp_param_to_clean[alias] = clean
            bm.alias_to_clean.setdefault(alias, clean)
            bm.clean_keys.add(clean)
        for alias, value in _DICT_HARD_RE.findall(src):
            if alias not in bm.sp_param_to_clean:
                bm.sp_param_hardcoded[alias] = value
        for role in _SP_TOKEN_RE.findall(src):
            bm.implemented_roles.add(role)
    return bm


# --------------------------------------------------------------------------- #
# Frontend (TS feature) extractor
# --------------------------------------------------------------------------- #
_ZOD_FIELD_RE = re.compile(r"^\s*(\w+)\s*:\s*z\.", re.MULTILINE)
_MUT_RE = re.compile(r"use(Create|Update|Delete|Insert)\w*Mutation", re.IGNORECASE)


_IMPORT_RE = re.compile(r"""from\s+['"]([^'"]+)['"]""")
# Imports of the shared zod models / api adapters the copy-step rule must include.
_SHARED_IMPORT_RE = re.compile(r"(?:^|[@/])(models|adapters)/", re.IGNORECASE)


def _is_shared_import(spec: str) -> bool:
    return bool(_SHARED_IMPORT_RE.search(spec))


def _shared_fragment(spec: str) -> str | None:
    """The `models/...` or `adapters/...` tail of a shared import, alias/relative
    prefix stripped — e.g. `@/models/raw-material` -> `models/raw-material`."""
    m = _SHARED_IMPORT_RE.search(spec)
    return spec[m.start(1):] if m else None


def _shared_import_present(frag: str, slice_paths: list[str]) -> bool:
    """A shared import is satisfied if any slice file resolves it as a FILE
    (`.../models/raw-material.ts`) or as a DIRECTORY/index
    (`.../models/sales-summary/index.ts`). Matching on the path fragment — not the
    bare stem — avoids false 'missing' on directory-organized shared modules."""
    needle = "/" + frag
    return any((needle + "." in sp) or (needle + "/" in sp) for sp in slice_paths)
_ZOD_OBJ_OPEN_RE = re.compile(r"(\w+)\s*=\s*z\.object\(")
_ZOD_KEY_RE = re.compile(r"^\s*(\w+)\s*:\s*z\.")
_GRID_FIELD_RE = re.compile(r"""\bfield\s*:\s*['"](\w+)['"]""")


def _classify_schema(name: str) -> str:
    low = name.lower()
    if any(t in low for t in ("queryparam", "params", "query", "filter", "search")):
        return "filter"
    if any(t in low for t in ("item", "row", "detail")) and "pagination" not in low:
        return "column"
    return "other"


def extract_frontend_feature(files: list[Path], anchor_dir: Path | None = None) -> FrontendFeature:
    ff = FrontendFeature(feature_dir=str(anchor_dir) if anchor_dir else "")
    for ts in files:
        ff.files.append(ts.name)
        src = ts.read_text(encoding="utf-8", errors="replace")
        for fld in _ZOD_FIELD_RE.findall(src):
            ff.zod_fields.add(fld)
        # Stateful pass: classify each named zod schema's fields.
        current: str | None = None
        for line in src.splitlines():
            m = _ZOD_OBJ_OPEN_RE.search(line)
            if m:
                current = _classify_schema(m.group(1))
                continue
            if current:
                km = _ZOD_KEY_RE.match(line)
                if km:
                    if current == "filter":
                        ff.filter_fields.add(km.group(1))
                    elif current == "column":
                        ff.column_fields.add(km.group(1))
                if "})" in line:
                    current = None
        # Grid column definitions (AG-Grid style field: '...').
        for fld in _GRID_FIELD_RE.findall(src):
            ff.column_fields.add(fld)
        for verb in _MUT_RE.findall(src):
            v = verb.lower()
            ff.mutations.add("create" if v in {"create", "insert"} else v)
    return ff


# --------------------------------------------------------------------------- #
# Resolver: screen_id -> spec file / backend module / frontend feature
# --------------------------------------------------------------------------- #
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def resolve_spec_file(spec_root: Path, screen_id: str) -> Path | None:
    key = _norm(screen_id)
    for f in sorted(spec_root.rglob("*.xlsx")):
        if _norm(f.stem).startswith(key) or key in _norm(f.stem):
            return f
    return None


def _is_module_dir(d: Path) -> bool:
    return d.is_dir() and bool(list(d.glob("*.py")))


def _alpha_digit_key(s: str) -> tuple[tuple[str, ...], str]:
    """Split a normalized id into (sorted-letters, digits-in-original-order).
    Letters are a multiset so a spec/code transposition like `PRG` vs dir `pgr`
    still matches, but digits keep their sequence so neither 003-vs-004 NOR the
    digit-anagram 010-vs-001 can collide."""
    n = _norm(s)
    letters = tuple(sorted(c for c in n if c.isalpha()))
    digits = "".join(c for c in n if c.isdigit())
    return letters, digits


def resolve_backend_module(backend_root: Path, screen_id: str) -> Path | None:
    """Match by screen-id code. Exact normalized match first, then a
    letters-multiset + digits-in-order match so a spec/code transposition like
    `PRG` vs dir `pgr` still binds `PDT_PRG_003M` -> `pdt_pgr_003m`, while the
    ordered digits keep 003≠004 and 010≠001 (digit anagrams) from colliding."""
    key = _norm(screen_id)            # e.g. pdtprg003m
    kkey = _alpha_digit_key(screen_id)
    transposed: Path | None = None
    for d in sorted(p for p in backend_root.rglob("*") if _is_module_dir(p)):
        nd = _norm(d.name)
        if nd == key:
            return d
        if transposed is None and _alpha_digit_key(d.name) == kkey:
            transposed = d
    return transposed


_SP_BASE_RE = re.compile(r"\b(str_\w+?)_(?:S|IU|I|U|D)\b")


def resolve_backend_by_sps(backend_root: Path, sp_names: list[str]) -> Path | None:
    """Fallback resolver (used when the screen-id code doesn't match a module):
    bind by stored-procedure names, weighting each SP by *specificity* (inverse
    document frequency). A generic SP shared by many modules counts for little; a
    unique SP that appears in only one module is decisive. This avoids the
    over-binding where several 진행현황 screens collapse onto one module via a
    common procedure."""
    bases = {re.sub(r"_(?:S|IU|I|U|D)(?:_\d+)?$", "", n.rsplit(".", 1)[-1]) for n in sp_names}
    bases = {b for b in bases if b}
    if not bases:
        return None
    contains: dict[Path, set[str]] = {}
    for d in sorted(p for p in backend_root.rglob("*") if _is_module_dir(p)):
        text = "\n".join(p.read_text(encoding="utf-8", errors="replace") for p in d.glob("*.py"))
        hit = {b for b in bases if b in text}
        if hit:
            contains[d] = hit
    if not contains:
        return None
    df = {b: sum(1 for s in contains.values() if b in s) for b in bases}
    best, best_score = None, 0.0
    for d, hit in contains.items():
        score = sum(1.0 / df[b] for b in hit if df.get(b))
        if score > best_score:
            best, best_score = d, score
    return best


def resolve_frontend_files(
    frontend_root: Path,
    backend: BackendModule,
    screen_id: str,
    feature_dir: Path | None = None,
) -> tuple[list[Path], Path | None, set[str]]:
    """A screen's frontend files are spread across feature/<slug>, models/,
    hooks/<slug>, adapters/. Seed from files that reference the backend's stored
    procedures (or screen id), then follow their imports to pull in the model and
    hook files (where the zod fields and mutations live).

    When ``feature_dir`` is given (an explicit screen→frontend binding), the seed
    is anchored deterministically to every file under that directory instead of
    discovered by SP/id heuristics — imports are still followed across
    ``frontend_root`` so shared ``src/models/`` is pulled in.

    Returns ``(files, anchor_dir, missing_shared)`` where ``missing_shared`` is the
    set of shared ``@/models`` / ``@/adapters`` import specifiers a seed file
    references but that are absent from the slice — the precise signal that the
    copy-step rule was violated (only ``feature-*/`` was copied)."""
    sp_bases: set[str] = set()
    for py in Path(backend.module_dir).rglob("*.py"):
        for m in re.findall(r"(str_\w+?)_(?:IU|S|D)", py.read_text(encoding="utf-8", errors="replace")):
            sp_bases.add(m)

    all_ts = list(frontend_root.rglob("*.ts")) + list(frontend_root.rglob("*.tsx"))
    by_stem: dict[str, list[Path]] = {}
    for ts in all_ts:
        by_stem.setdefault(ts.stem, []).append(ts)

    seed: list[Path] = []
    anchor_dir: Path | None = None
    if feature_dir is not None:
        anchor_dir = feature_dir
        seed = sorted(p for p in (list(feature_dir.rglob("*.ts")) + list(feature_dir.rglob("*.tsx"))))
    else:
        for ts in all_ts:
            src = ts.read_text(encoding="utf-8", errors="replace")
            if any(b in src for b in sp_bases) or _norm(screen_id) in _norm(src):
                seed.append(ts)
                if anchor_dir is None and ts.name.endswith("Controller.tsx"):
                    anchor_dir = ts.parent

    slice_paths = [p.as_posix() for p in all_ts]
    cluster: dict[Path, None] = {p: None for p in seed}
    missing_shared: set[str] = set()
    for ts in seed:                       # follow imports to model/hook files
        src = ts.read_text(encoding="utf-8", errors="replace")
        for imp in _IMPORT_RE.findall(src):
            stem = imp.rsplit("/", 1)[-1]
            targets = by_stem.get(stem, [])
            for target in targets:
                cluster.setdefault(target, None)
            if targets or not _is_shared_import(imp):
                continue
            # No file-stem match: resolve the shared import by path fragment so a
            # directory/index import still counts. Pull its index files into the
            # cluster; only flag missing when nothing in the slice resolves it —
            # that is the genuine copy-step violation (feature-only copy).
            frag = _shared_fragment(imp)
            if frag and _shared_import_present(frag, slice_paths):
                for p in all_ts:
                    if ("/" + frag + "/") in p.as_posix():
                        cluster.setdefault(p, None)
            else:
                missing_shared.add(imp)
    return list(cluster), anchor_dir, missing_shared
