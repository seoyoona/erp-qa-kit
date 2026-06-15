from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import RawExtraction


name = "xlsx_text"
supported_extensions = frozenset({".xlsx", ".xlsm"})
tier = "REQUIRED"


def available() -> bool:
    return True


def extract(path: Path) -> RawExtraction:
    workbook = load_workbook(path, data_only=True, read_only=False)
    records: list[dict[str, Any]] = []
    merged_ranges: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        for merged in sheet.merged_cells.ranges:
            merged_ranges.setdefault(sheet.title, []).append(str(merged))
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                records.append(
                    {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": str(cell.value).strip(),
                    }
                )
    notes = [
        f"{sheet}: merged ranges {', '.join(ranges)}"
        for sheet, ranges in sorted(merged_ranges.items())
        if ranges
    ]
    return RawExtraction(
        source_path=path.as_posix(),
        kind="tabular",
        records=records,
        provenance_hint="medium",
        notes=notes,
        adapter=name,
    )

