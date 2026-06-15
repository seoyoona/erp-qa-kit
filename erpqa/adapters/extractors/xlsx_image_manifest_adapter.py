from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import RawExtraction


name = "xlsx_image_manifest"
supported_extensions = frozenset({".xlsx", ".xlsm"})
tier = "REQUIRED"


def available() -> bool:
    return True


def extract(path: Path) -> RawExtraction:
    workbook = load_workbook(path, data_only=True, read_only=False)
    records: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for index, image in enumerate(getattr(sheet, "_images", []), start=1):
            anchor = getattr(image, "anchor", None)
            anchor_cell = "unknown"
            if hasattr(anchor, "_from"):
                marker = anchor._from
                anchor_cell = f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
            records.append(
                {
                    "sheet": sheet.title,
                    "image_index": index,
                    "anchor_cell": anchor_cell,
                    "width": getattr(image, "width", None),
                    "height": getattr(image, "height", None),
                    "text": None,
                    "ocr_performed": False,
                }
            )
    return RawExtraction(
        source_path=path.as_posix(),
        kind="image_manifest",
        records=records,
        provenance_hint="low",
        notes=["embedded images are manifested only; OCR is not performed"],
        adapter=name,
    )

